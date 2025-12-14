#!/usr/bin/env python3
"""
Excel Analyzer
Analyzes Excel files to extract structure, formulas, and calculated values.
Saves results to a .txt file.
"""

import sys
import os
import openpyxl
from openpyxl.utils import get_column_letter


def analyze_excel(filepath):
    """Analyze an Excel file and return the analysis as a string."""

    # Load workbook twice: once for formulas, once for calculated values
    wb_formulas = openpyxl.load_workbook(filepath)
    wb_values = openpyxl.load_workbook(filepath, data_only=True)

    lines = []
    lines.append("=" * 60)
    lines.append(f"EXCEL ANALYSIS: {filepath}")
    lines.append("=" * 60)
    lines.append(f"\nSheet names: {wb_formulas.sheetnames}")

    # Analyze each sheet
    for sheet_name in wb_formulas.sheetnames:
        ws_formulas = wb_formulas[sheet_name]
        ws_values = wb_values[sheet_name]

        lines.append(f"\n{'=' * 60}")
        lines.append(f"SHEET: {sheet_name}")
        lines.append("=" * 60)
        lines.append(f"Dimensions: {ws_formulas.dimensions}")
        lines.append(f"Max Row: {ws_formulas.max_row}, Max Column: {ws_formulas.max_column}")

        lines.append("\nCell Contents:")
        lines.append("-" * 40)

        for row_idx in range(1, ws_formulas.max_row + 1):
            row_data = []
            for col_idx in range(1, ws_formulas.max_column + 1):
                cell_f = ws_formulas.cell(row=row_idx, column=col_idx)
                cell_v = ws_values.cell(row=row_idx, column=col_idx)

                if cell_f.value is not None:
                    cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
                    formula_val = cell_f.value
                    calc_val = cell_v.value

                    # Check if it's a formula
                    if isinstance(formula_val, str) and formula_val.startswith("="):
                        row_data.append(f"{cell_ref}: {formula_val} -> {calc_val}")
                    else:
                        row_data.append(f"{cell_ref}: {repr(formula_val)}")

            if row_data:
                lines.append(f"  Row {row_idx}:")
                for item in row_data:
                    lines.append(f"    {item}")

        # Summary of formulas
        lines.append("\nFormulas Found:")
        lines.append("-" * 40)
        formula_count = 0
        for row in ws_formulas.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                    calc_val = ws_values[cell_ref].value
                    lines.append(f"  {cell_ref}: {cell.value}")
                    lines.append(f"         = {calc_val}")
                    formula_count += 1

        if formula_count == 0:
            lines.append("  (No formulas found in this sheet)")

    lines.append("\n" + "=" * 60)
    lines.append("ANALYSIS COMPLETE")
    lines.append("=" * 60)

    return "\n".join(lines)


def main():
    if len(sys.argv) < 2:
        print("Usage: python excel-analyzer.py <excel_file.xlsx>")
        print("\nExample:")
        print("  python excel-analyzer.py excel-calculation-engine.xlsx")
        print("  python excel-analyzer.py excel-lookup-engine.xlsx")
        sys.exit(1)

    filepath = sys.argv[1]

    # Generate output filename (same name but .txt extension)
    base_name = os.path.splitext(filepath)[0]
    output_file = f"{base_name}.txt"

    try:
        result = analyze_excel(filepath)

        with open(output_file, "w", encoding="utf-8") as f:
            f.write(result)

        print(f"Analysis saved to: {output_file}")

    except FileNotFoundError:
        print(f"Error: File '{filepath}' not found.")
        sys.exit(1)
    except Exception as e:
        print(f"Error analyzing file: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
